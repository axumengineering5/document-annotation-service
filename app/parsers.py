"""
parsers.py — extract raw text from PDF and spreadsheet files

Dependencies:
  pip install pdfplumber openpyxl
"""

import io


def parse_document(filename: str, content: bytes) -> str:
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext == "pdf":
        return _parse_pdf(content)
    if ext in ("xlsx", "xls", "csv"):
        return _parse_spreadsheet(filename, content)
    # Fallback: treat as plain text
    return content.decode("utf-8", errors="replace")


def _parse_pdf(content: bytes) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            return "\n".join(
                page.extract_text() or "" for page in pdf.pages
            )
    except ImportError:
        return "[pdfplumber not installed — install it to parse PDFs]"


def _parse_spreadsheet(filename: str, content: bytes) -> str:
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext == "csv":
        return content.decode("utf-8", errors="replace")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"=== Sheet: {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                lines.append("\t".join(str(c) if c is not None else "" for c in row))
        return "\n".join(lines)
    except ImportError:
        return "[openpyxl not installed — install it to parse XLSX files]"
